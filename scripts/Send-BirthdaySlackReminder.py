import os
from datetime import datetime
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook


EXCEL_PATH = "data/birthdays.xlsx"
WORKSHEET_NAME = "Birthdays"


def get_today_ist():
    """Return current date/time in India Standard Time."""
    return datetime.now(ZoneInfo("Asia/Kolkata"))


def read_birthdays():
    """Read birthday records from Excel."""
    workbook = load_workbook(
        EXCEL_PATH,
        data_only=True
    )

    worksheet = workbook[WORKSHEET_NAME]

    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]

    required_columns = {
        "Name",
        "Birthday",
        "SlackUserId"
    }

    missing_columns = required_columns - set(headers)

    if missing_columns:
        raise ValueError(
            f"Missing Excel columns: {', '.join(sorted(missing_columns))}"
        )

    records = []

    for row in rows[1:]:
        record = dict(zip(headers, row))
        records.append(record)

    return records


def find_todays_birthdays(records, today):
    """Find people whose birthday matches today's month and day."""
    birthdays = []

    for person in records:
        name = str(person.get("Name") or "").strip()
        birthday_value = person.get("Birthday")

        if not name or not birthday_value:
            continue

        try:
            if isinstance(birthday_value, datetime):
                birthday = birthday_value
            else:
                birthday = datetime.fromisoformat(
                    str(birthday_value).strip()
                )

            print(
                f"Checking {name}: "
                f"Birthday={birthday.strftime('%Y-%m-%d')}"
            )

            # Year is intentionally ignored.
            if (
                birthday.month == today.month
                and birthday.day == today.day
            ):
                birthdays.append(person)

        except (ValueError, TypeError):
            print(
                f"WARNING: Invalid Birthday value "
                f"for {name}: {birthday_value}"
            )

    return birthdays


def send_slack_message(person, webhook_url):
    """Send birthday notification to Slack."""
    name = str(person.get("Name") or "").strip()
    slack_user_id = str(
        person.get("SlackUserId") or ""
    ).strip()

    if slack_user_id:
        message = (
            f"🎂 Happy Birthday <@{slack_user_id}>! "
            f"Wishing you a fantastic day! 🎉"
        )
    else:
        message = (
            f"🎂 Today is {name}'s birthday! "
            f"Please wish them a Happy Birthday! 🎉"
        )

    payload = {
        "text": message
    }

    response = requests.post(
        webhook_url,
        json=payload,
        timeout=30
    )

    response.raise_for_status()

    print(
        f"Slack reminder sent successfully for {name}."
    )


def main():
    print("==========================================")

    today = get_today_ist()

    print(f"India Date/Time : {today}")
    print(
        f"India Date      : "
        f"{today.strftime('%Y-%m-%d')}"
    )

    print("==========================================")

    webhook_url = os.getenv("SLACK_WEBHOOK_URL")

    if not webhook_url:
        raise RuntimeError(
            "SLACK_WEBHOOK_URL environment variable is not set."
        )

    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(
            f"Excel file not found: {EXCEL_PATH}"
        )

    records = read_birthdays()

    birthdays_today = find_todays_birthdays(
        records,
        today
    )

    if not birthdays_today:
        print("No birthdays today.")
        return

    for person in birthdays_today:
        name = str(person.get("Name") or "").strip()

        print(f"Sending Slack reminder for {name}...")

        send_slack_message(
            person,
            webhook_url
        )


if __name__ == "__main__":
    main()